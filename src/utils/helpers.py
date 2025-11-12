import logging
import asyncio
import openpyxl
import base64
import requests
import secrets
import re
import uuid

from aiogram.types import LabeledPrice
from transliterate import translit
from aiocryptopay import AioCryptoPay, Networks
from aiogram.utils.deep_linking import create_start_link
from aiogram.types import Message
from typing import List
from openpyxl.styles import Font, Alignment, PatternFill
from aiogram.types import Message, CallbackQuery
from aiogram.fsm.context import FSMContext
from aiogram import Bot
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import datetime
from dateutil.relativedelta import relativedelta

from src.config import settings
from src.database.models import User
from src.database.repositories import PayRepository, UserRepository
from src.keyboards.user_keyboards import user_menu

logger = logging.getLogger(__name__)

crypto = AioCryptoPay(token=settings.CRYPTO_PAY_TOKEN, network=Networks.TEST_NET)


async def safe_answer(message: Message, text: str, reply_markup=None, **kwargs):
    try:
        return await message.answer(
            text=text, 
            reply_markup=reply_markup, 
            **kwargs
        )
    except Exception as e:
        logger.error(f"Ошибка при отправке сообщения: {e}")
        try:
            return await message.answer(text=text, **kwargs)
        except Exception as e2:
            logger.error(f"Критическая ошибка отправки сообщения: {e2}")
            return None


async def try_edit_callback(callback: CallbackQuery, text: str, reply_markup=None, **kwargs):
    try:
        return await callback.message.edit_text(
            text=text,
            reply_markup=reply_markup,
            **kwargs
        )
    except Exception as e:
        logger.error(f"Ошибка при редактировании сообщения: {e}")
        try:
            await callback.message.delete()
        except:
            pass
        return await callback.message.answer(
            text=text,
            reply_markup=reply_markup,
            **kwargs
        )


async def delete_state_message(state: FSMContext, message: Message) -> dict:
    state_info = await state.get_data()
    try:
        await message.bot.delete_messages(
            chat_id=message.from_user.id,
            message_ids=[
                state_info.get("mes_del"),
                message.message_id
            ]
        )
    except:
        logging.error("Error delete message in state")
    return state_info


async def answer_user_message(state_info: dict, bot: Bot, user_id: int):
    answer_users = []
    not_answer_users = []
    answer_btn = state_info.get("result_btn")
    res_type = state_info.get("res_type")
    text = state_info.get("text")
    photo = state_info.get("photo")
    
    try:
        if res_type == "text":
            await bot.send_message(
                chat_id=user_id,
                text=text,
                parse_mode="HTML",
                reply_markup=answer_btn
            )
        else:
            await bot.send_photo(
                photo=photo,
                chat_id=user_id,
                caption=text,
                parse_mode="HTML",
                reply_markup=answer_btn
            )
        answer_users.append(user_id)
    except:
        not_answer_users.append(user_id)
    await asyncio.sleep(0.2)


def export_users_to_excel(users: List[User]):
    filename = f'users_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Пользователи"

    headers = [
        " ", "Айди пользователя", "Юзернейм пользователя", "Фулл нейм пользователя", "Дата регистрации",
        "Приглашен ли", "Баланс", "Реферальный баланс", "Активна ли подписка", "Ключ"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    for row_num, row_data in enumerate(users, 2):
        
        row_values = [
            row_num-1,
            row_data.user_id,
            row_data.username if row_data.username else "Отсутствует",
            row_data.full_name,
            row_data.created_at.strftime('%d.%m.%Y'),
            "Да" if row_data.referrer_id else "Нет",
            row_data.main_balance,
            row_data.referral_balance,
            "Да" if row_data.is_active else "Нет",
            row_data.vpn_key if row_data.vpn_key else "Отсутствует"
        ]

        for col_num, cell_value in enumerate(row_values, 1):
            ws.cell(row=row_num, column=col_num).value = cell_value

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = (max_length + 2) * 1.2

    wb.save(filename)
    return filename


async def fix_base64_padding(payload):
    missing_padding = len(payload) % 4
    if missing_padding:
        payload += '=' * (4 - missing_padding)
    return payload


async def decode_payload(payload: str) -> str:
    payload = await fix_base64_padding(payload)
    return base64.b64decode(payload).decode("utf-8")

async def get_reflink(user_id: str, bot: Bot):
    start_link = await create_start_link(bot, str(user_id), encode=True)
    return start_link


async def create_invoice_crypto_pay(callback: CallbackQuery, amount: int, pay_id: int):
    try:
        invoice = await crypto.create_invoice(
            asset="USDT",
            amount=str(amount),
            description=f"💰 Пополнение баланса на {amount} USDT",
            payload=f"{callback.message.chat.id}:{int(amount)}:{str(pay_id)}",
        )
    except Exception as e:
        logging.warning(f"Error create invoice: {e}")
        return None
    if hasattr(invoice, "bot_invoice_url"):
        url=invoice.bot_invoice_url
        return url
    else:
        logging.error("not found error in create invoice")
        return None
        

async def pay_process(session: AsyncSession, pay_id: int, amount: int, bot: Bot):
    pay_info = await PayRepository.get_pay(
        async_session=session,
        pay_id=pay_id
    )
 
    user_info = await UserRepository.give_user(
        async_session=session,
        user_id=pay_info.user_id
    )
    if user_info.referrer_id:
        try:
            txt = "✅ Приглашенный вами друг пополнил баланс, вам начислено 50 бонусных рублей"
            await bot.send_message(
                chat_id=user_info.referrer_id,
                text=txt,
                parse_mode="HTML"
            )
        except:
            logging.error("error send message")
        await UserRepository.plus_reffered_balance(
            async_session=session,
            user_id=user_info.referrer_id,
            amount=50
        )
    new_balance = await UserRepository.plus_balance(
        async_session=session,
        user_id=pay_info.user_id,
        amount=amount
    )
    try:
        txt = f"✅ Вы успешно оплатили подписку\n\nВаш текущий баланс: {new_balance}"
        btn = user_menu
        await bot.send_message(
            chat_id=pay_info.user_id,
            text=txt,
            reply_markup=btn,
            parse_mode="HTML"
        )
    except:
        logging.error("error send message")


async def give_me_key(full_name: str, ):
    my_date = datetime.now() + relativedelta(years=20)
    expireAt = my_date.isoformat() + 'Z'
    url = settings.VPN_BASE_URL
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {settings.VPN_KEY}"
    }
    
    full_name = translit(full_name, 'ru', reversed=True) if any(c in 'абвгдеёжзийклмнопрстуфхцчшщъыьэюя' for c in full_name.lower()) else full_name
    full_name = re.sub(r'[^a-zA-Z0-9_-]', '_', full_name)  
    full_name = re.sub(r'_+', '_', full_name).strip('_') 
    if len(full_name) < 3:
        full_name = f"user_{secrets.token_hex(4)}" 

    my_date = datetime.now() + relativedelta(years=20)
    expireAt = my_date.isoformat() + 'Z'
    data = {
        "username": full_name,
        "status": "ACTIVE",
        "shortUuid": secrets.token_hex(4),  # Короткий уникальный ID (8 chars)
        "trojanPassword": secrets.token_hex(8),  # Рандомный пароль >=8 chars
        "vlessUuid": str(uuid.uuid4()),  # Валидный UUID
        "ssPassword": secrets.token_hex(8),  # Рандомный пароль >=8 chars
        "trafficLimitBytes": settings.VPN_GB,
        "trafficLimitStrategy": "NO_RESET",
        "expireAt": expireAt,
        "description": "My VPN user",
        "hwidDeviceLimit": 0,
        "activeInternalSquads": []  # Пустой массив
        # Опущены: createdAt, lastTrafficResetAt, uuid, tag, telegramId, email, externalSquadUuid — сервер сгенерирует
    }

    # Отправка POST-запроса
    response = requests.post(url, headers=headers, json=data)

    # Проверка ответа
    if response.status_code == 200 or response.status_code == 201:  # Успех (created)
        logging.info("Успех! Ответ от API:")
        logging.info(response.json())  # Здесь будут ключи: uuid, vlessUuid, trojanPassword и т.д.
    else:
        logging.info(f"Ошибка: {response.status_code}")
        logging.info(response.text)  # Для отладки

    # Пример: извлечение ключей из ответа
    if response.status_code in (200, 201):
        result = response.json().get("response", {})
        vless_uuid = result.get("vlessUuid")
        trojan_password = result.get("trojanPassword")
        subscription_url = result.get("subscriptionUrl")
        logging.info(f"VLESS UUID: {vless_uuid}")
        logging.info(f"Trojan Password: {trojan_password}")
        logging.infos(f"Subscription URL: {subscription_url}")


async def create_invoice(message: Message):
    one_star = LabeledPrice(label='Доступ к VPN на 1 месяц', amount=10)
    await message.bot.send_invoice(
        chat_id=message.chat.id,
        title="Оплата VPN-доступа",
        description="Получи ключ для VPN на 20 лет за 10 Stars",
        provider_token="", 
        currency="XTR", 
        photo_url="https://example.com/vpn_photo.jpg", 
        photo_width=800,  
        photo_height=600,
        photo_size=100000,  
        is_flexible=False, 
        prices=[one_star],
        start_parameter="vpn-access", 
        payload="vpn:10_stars" 
    )